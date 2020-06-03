#13_excel.py

# Let's try the same thing with an excel sheet.
# What could go wrong?
# Excel sheets have fancier formatting.
# Next time I'll tell you about nice things you can do with Excel,
# But today, I'll just bring up the problem you will always have to deal with:
# They can store multiple sheets.

# To load data, you need load_workbook.
# That opens the file for you.
from openpyxl import load_workbook

# I converted the CSV file into an excel workbook,
# and I also separated the data into two sheets,
# so that one contains the names, and one contains the locations.
workbook = load_workbook("subways.xlsx")
# If you don't know the sheets, you can get them from the sheetnames field.
print("Sheets: %s" % (workbook.sheetnames))
# You use get_sheet_by_name to get the individual sheets
name_sheet = workbook.get_sheet_by_name("Index")
location_sheet = workbook.get_sheet_by_name("Locations")

# Let's see what the classes have to offer.
print("A workbook has attributes: %s" % (dir(workbook)))
print("A sheet has attributes: %s" % (dir(name_sheet)))

# This time, let's keep two dictionaries, one that maps IDs to names,
# and another that maps IDs to locations.
# The Excel module has some nice features that resemble
# how you would use a spreadsheet program.
# The column names are actually letters,
# and you index them by letters.
# And just like you can select a rectangle of cells,
# you can iterate over them.
# For the index, we would want to select columns B and C,
# over all rows except for the first one.
name_lookup = {}
# We use max_row to get all the way to the bottom.
name_cells = name_sheet["B2" : "C" + str(name_sheet.max_row)]
# But at the first level, you iterate over rows again:
for row in name_cells:
	objectid, name = map(lambda cell: cell.value, row)
	name_lookup[objectid] = name

# We use the same rows for the locations, but columns A to B.
location_lookup = {}
# We use max_row to get all the way to the bottom.
location_cells = location_sheet["A2" : "B" + str(location_sheet.max_row)]
for row in location_cells:
	objectid, location_str = map(lambda cell: cell.value, row)
	# Let's convert the location to numbers again
	location = tuple(map(float, location_str[7 : -1].split(" ")))
	location_lookup[objectid] = location

# Asking for your location again, and calculating and saving the distances.
latitude = None
while latitude is None:
	latitude_str = input("Please enter the latitude: ")
	try:
		latitude = float(latitude_str)
	except ValueError:
		print("%s is not a real number" % latitude_str)
longitude = None
while longitude is None:
	longitude_str = input("Please enter the longitude: ")
	try:
		longitude = float(longitude_str)
	except ValueError:
		print("%s is not a real number" % longitude_str)

from distance import earth_dist

min_dist = None
closest_station = None
closest_location = None
# This time, let's look up the stations by object ID.
distance_lookup = {}
for (object_id, (station_longitude, station_latitude)) in \
    location_lookup.items():
	dist = earth_dist(latitude, longitude, \
			  station_latitude, station_longitude)
	# To get the name, we need the first sheet's data.
	name = name_lookup[object_id]
	if min_dist is None or dist < min_dist:
		min_dist = dist
		closest_station = name
		closest_location = (station_longitude, station_latitude)
	distance_lookup[object_id] = dist

print("The closest station is %s, at (%f, %f), %f kilometres away" %
      (closest_station, closest_location[1], closest_location[0], min_dist))

# Now, let's save to an Excel sheet this time.
# In fact, let's put it in a workbook with the old data.
distance_sheet = workbook.create_sheet(title = "distances")
# Let's write the column headers.
distance_sheet["A1"] = "name"
distance_sheet["B1"] = "distance"
# Let's get the name for each object ID from the first sheet.
# So we need both the object ID and the name.
objectids = name_sheet["B2" : "C" + str(name_sheet.max_row)]

for id_row in objectids:
	# This is a row, which could have multiple cells,
	# so it acts like an indexed collection, and we know
	# what the two cells are.
	objectid = id_row[0].value
	name = id_row[1].value
	# Put the new row in the next open one.
	distance_sheet.append([name, distance_lookup[objectid]])
# We finished writing. Now save, but in a new file.
workbook.save("extended_subways.xlsx")
